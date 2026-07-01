"""Excel export utilities (writes a brand-new cleaned file, never overwrites input)."""
from __future__ import annotations

import os

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import config

def build_output_path(input_path: str) -> str:
    """Return Cleaned_Data_<OriginalFileName>.xlsx in the input's directory."""
    directory = os.path.dirname(os.path.abspath(input_path))
    base = os.path.basename(input_path)
    name, _ = os.path.splitext(base)
    output_name = f"{config.OUTPUT_PREFIX}{name}.xlsx"
    return os.path.join(directory, output_name)

def export_to_excel(df: pd.DataFrame, input_path: str) -> str:
    """Write the cleaned DataFrame to a new .xlsx file and return its path.

    Restores the original Filename hyperlinks (pandas drops them on read) so
    the cleaned file's filenames remain clickable links to the creatives.
    """
    output_path = build_output_path(input_path)

    # Defensive: never overwrite the original input file.
    if os.path.abspath(output_path) == os.path.abspath(input_path):
        raise IOError("Refusing to overwrite the original input file.")

    export_df = df.copy()
    links = df.attrs.get("filename_links", {})

    with pd.ExcelWriter(
        output_path, engine="openpyxl", datetime_format="YYYY-MM-DD"
    ) as writer:
        export_df.to_excel(writer, index=False, sheet_name="Cleaned Data")
        worksheet = writer.sheets["Cleaned Data"]

        # Light auto-fit for readability (data values themselves are unchanged).
        for i, column in enumerate(export_df.columns, start=1):
            values = export_df[column].astype(str)
            max_len = max([len(str(column))] + [len(v) for v in values])
            worksheet.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)

        # Restore clickable Filename hyperlinks.
        if links:
            filename_idx = config.OUTPUT_COLUMNS.index("Filename") + 1  # 1-based
            link_font = Font(color="0563C1", underline="single")
            for row in range(2, worksheet.max_row + 1):  # row 1 = header
                cell = worksheet.cell(row=row, column=filename_idx)
                if cell.value is None:
                    continue
                url = links.get(str(cell.value).strip())
                if url:
                    cell.hyperlink = url
                    cell.font = link_font

    return output_path