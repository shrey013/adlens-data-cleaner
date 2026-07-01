"""Core cleaning engine for raw AdClarity Excel exports."""
from __future__ import annotations

import os
from typing import Callable, Dict, List, Optional

import pandas as pd

import config
import filters
from utils import column_letter_to_index, parse_date, parse_impressions

class CleaningError(Exception):
    """Raised when the cleaning process cannot complete (friendly message)."""

# Signature: (step_name, status_message, percentage_0_to_1)
ProgressCallback = Callable[[str, str, float], None]

class AdClarityCleaner:
    """Cleans a raw AdClarity Excel export into a normalised dataset.

    Self-contained so future modules (AI classification, OCR, dashboards,
    etc.) can consume the cleaned DataFrame without modifying this engine.
    """

    def __init__(self, progress_callback: Optional[ProgressCallback] = None) -> None:
        self._progress: ProgressCallback = progress_callback or (lambda s, m, p: None)

    # ---- internal helpers -------------------------------------------------

    def _emit(self, step: str, status: str, pct: float) -> None:
        self._progress(step, status, pct)

    def _detect_header_row(self, raw: pd.DataFrame) -> int:
        """Find the row containing all required header labels.

        Searches the entire sheet; never assumes a fixed row number.
        """
        required = set(config.REQUIRED_HEADERS)
        for idx in range(len(raw)):
            row_values = {
                str(cell).strip().lower()
                for cell in raw.iloc[idx].tolist()
                if cell is not None and str(cell).strip() != ""
            }
            if required.issubset(row_values):
                return idx
        raise CleaningError(
            "Missing Required Columns: could not locate the AdClarity header "
            "row (Filename, Impressions, First Seen, Last Seen)."
        )

    def _resolve_columns(self, header_values: List[object]) -> Dict[str, int]:
        """Map each output column to a source index.

        Primary: match by header name (robust to layout shifts).
        Fallback: fixed column letters D, N, R, S from the spec.
        """
        lowered = [
            str(h).strip().lower() if h is not None else "" for h in header_values
        ]
        mapping: Dict[str, int] = {}
        for canonical, target in zip(config.OUTPUT_COLUMNS, config.REQUIRED_HEADERS):
            if target in lowered:
                mapping[canonical] = lowered.index(target)
            else:
                letter = config.FALLBACK_COLUMN_LETTERS[canonical]
                mapping[canonical] = column_letter_to_index(letter)
        return mapping

    def _extract_filename_links(self, file_path: str) -> Dict[str, str]:
        """Read embedded Filename hyperlinks from the raw workbook.

        pandas strips hyperlinks on read, so we grab them separately with
        openpyxl and return a {filename_text: url} map. Used both for
        URL-based de-duplication and to restore clickable links on export.
        Returns an empty dict for .xls or files without links (not fatal).
        """
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=False)
        except Exception:  # noqa: BLE001 — old .xls or unreadable hyperlinks
            return {}

        required = set(config.REQUIRED_HEADERS)
        links: Dict[str, str] = {}

        for ws in wb.worksheets:
            header_row = None
            filename_col = None
            for r in range(1, min(ws.max_row, 50) + 1):
                row_map = {}
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(r, c).value
                    if v is not None and str(v).strip():
                        row_map[str(v).strip().lower()] = c
                if required.issubset(row_map.keys()):
                    header_row = r
                    filename_col = row_map["filename"]
                    break

            if header_row is None:
                continue

            for r in range(header_row + 1, ws.max_row + 1):
                cell = ws.cell(r, filename_col)
                if cell.value is None:
                    continue
                if cell.hyperlink and cell.hyperlink.target:
                    links[str(cell.value).strip()] = cell.hyperlink.target

            if links:
                break

        wb.close()
        return links

    def _remove_duplicates(self, df: pd.DataFrame, links: Dict[str, str]) -> pd.DataFrame:
        """Remove duplicate creatives, keeping the first occurrence.

        Basis (most accurate first):
          • If creative URLs are available, dedupe by URL — two rows pointing
            to the same creative are the same ad even if filenames differ.
          • Otherwise fall back to exact Filename match (per the spec).

        To use STRICT filename-only de-duplication instead, replace the body
        of this method with:
            return df.drop_duplicates(subset=["Filename"], keep="first").copy()
        """
        if links:
            tmp = df.copy()
            tmp["_url"] = tmp["Filename"].map(
                lambda f: links.get(str(f).strip(), str(f).strip())
            )
            tmp = tmp.drop_duplicates(subset=["_url"], keep="first")
            return tmp.drop(columns="_url").copy()
        return df.drop_duplicates(subset=["Filename"], keep="first").copy()

    # ---- public API -------------------------------------------------------

    def clean(
        self,
        file_path: str,
        target_year: int,
        target_month: int,
        min_impression: int,
    ) -> pd.DataFrame:
        """Run the full cleaning pipeline and return the cleaned DataFrame.

        The returned DataFrame carries the Filename→URL map in
        `df.attrs["filename_links"]` so the exporter can restore clickable
        hyperlinks.
        """
        if not file_path or not os.path.isfile(file_path):
            raise CleaningError("Invalid Excel: the selected file does not exist.")

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in config.ALLOWED_EXTENSIONS:
            raise CleaningError("Invalid Excel: only .xlsx and .xls files are supported.")

        # STEP 1 — Read Excel.
        self._emit("Reading Excel", "Loaded Excel", 0.05)
        try:
            raw = pd.read_excel(file_path, header=None, dtype=object)
        except Exception as exc:  # noqa: BLE001
            raise CleaningError("Corrupted File: unable to read the Excel file.") from exc

        if raw is None or raw.empty:
            raise CleaningError("Invalid Excel: the file contains no data.")

        # Extract creative hyperlinks up front (used for dedupe + export).
        filename_links = self._extract_filename_links(file_path)

        # STEP 2 — Detect header row dynamically.
        self._emit("Finding Header", "Header Found", 0.15)
        header_idx = self._detect_header_row(raw)
        header_values = raw.iloc[header_idx].tolist()
        column_map = self._resolve_columns(header_values)

        # STEP 3 — Drop metadata above header.
        data = raw.iloc[header_idx + 1:].reset_index(drop=True)
        if data.empty:
            raise CleaningError("No Matching Creatives: no data rows below the header.")

        # STEP 4 — Keep only required columns.
        if max(column_map.values()) >= data.shape[1]:
            raise CleaningError(
                "Missing Required Columns: expected columns are not present in the sheet."
            )

        self._emit("Cleaning Data", "Reading Columns", 0.25)
        cleaned = pd.DataFrame(
            {col: data.iloc[:, column_map[col]].values for col in config.OUTPUT_COLUMNS}
        )

        # STEP 5 — Columns already renamed via the dict keys above.

        # STEP 6 — Remove fully empty rows.
        cleaned = cleaned.dropna(how="all")

        # STEP 7 — Trim spaces and drop blank filenames.
        cleaned["Filename"] = cleaned["Filename"].apply(
            lambda v: str(v).strip() if v is not None and str(v).strip() != "" else None
        )
        cleaned = cleaned[cleaned["Filename"].notna()]
        if cleaned.empty:
            raise CleaningError("No Matching Creatives: no usable rows after cleaning.")

        # STEP 8 — Convert impressions.
        self._emit("Converting Impressions", "Converting Impressions", 0.40)
        cleaned["Impressions"] = cleaned["Impressions"].apply(parse_impressions)

        # STEP 9 — Convert dates (drop rows with invalid dates).
        self._emit("Converting Dates", "Converting Dates", 0.55)
        cleaned["First Seen"] = cleaned["First Seen"].apply(parse_date)
        cleaned["Last Seen"] = cleaned["Last Seen"].apply(parse_date)
        cleaned = cleaned[cleaned["First Seen"].notna() & cleaned["Last Seen"].notna()]
        if cleaned.empty:
            raise CleaningError("No Matching Creatives: no rows with valid dates.")

        # STEP 10 — Month filtering (active during month, NOT past month-end).
        self._emit("Filtering Month", "Filtering Month", 0.68)
        cleaned = filters.filter_by_month(cleaned, target_year, target_month)

        # STEP 11 — Impression filtering.
        self._emit("Filtering Impressions", "Filtering Impressions", 0.78)
        cleaned = filters.filter_by_impression(cleaned, min_impression)

        # STEP 12 — Duplicate removal (by creative URL, else by Filename).
        self._emit("Removing Duplicates", "Removing Duplicates", 0.86)
        cleaned = self._remove_duplicates(cleaned, filename_links)

        # STEP 13 — Sort descending by impressions.
        self._emit("Sorting", "Sorting", 0.92)
        cleaned = filters.sort_by_impressions(cleaned)

        if cleaned.empty:
            raise CleaningError(
                "No Matching Creatives: nothing matched the chosen month and impression filters."
            )

        # STEP 14 — Attach Filename→URL map for the exporter.
        result = cleaned[config.OUTPUT_COLUMNS].copy()
        result.attrs["filename_links"] = filename_links
        return result